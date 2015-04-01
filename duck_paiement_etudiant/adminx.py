# coding=utf-8
import datetime
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Sum, Count
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Style, Font, Border, Side, PatternFill, Alignment
import openpyxl.styles.borders
import openpyxl.styles.fills
import openpyxl.styles.colors
from openpyxl.writer.excel import save_virtual_workbook
from wkhtmltopdf.views import PDFTemplateView
from duck_paiement_etudiant.forms import PaiementBackofficeForm
from foad.models import AuditeurLibreApogee
from xadmin.layout import Layout, Fieldset, Container, Col
from django.forms import Media
from django.views.decorators.cache import never_cache
from django_apogee.models import InsAdmEtp
from duck_paiement_etudiant.models import AnneeUniPaiement, PaiementBackoffice, Banque, Bordereau, SettingEtapePaiement, \
    InsAdmEtpPaiement, PaiementAuditeurBackoffice
from xadmin.views import filter_hook, BaseAdminView
from xadmin.filters import NumberFieldListFilter

__author__ = 'paulguichon'
from xadmin import views
import xadmin


class GestionFinanciereAnnee(views.Dashboard):
    base_template = 'duck_paiement_etudiant/gestion_financiere_annee.html'
    widget_customiz = False
    site_title = 'Backoffice'

    def get_context(self):
        context = super(GestionFinanciereAnnee, self).get_context()
        context['years'] = AnneeUniPaiement.objects.all().order_by('-cod_anu')
        return context

    @filter_hook
    def get_breadcrumb(self):
        return [{'url': self.get_admin_url('index'), 'title': 'Accueil'},
                {'title': 'Gestion financière'}]

    @never_cache
    def get(self, request, *args, **kwargs):
        self.widgets = self.get_widgets()
        return self.template_response(self.base_template, self.get_context())
xadmin.site.register_view(r'^gestion_financiere/$', GestionFinanciereAnnee, 'gestion_financiere_annee')


class ListeImpayesAnnee(views.Dashboard):
    base_template = 'duck_paiement_etudiant/liste_impayes_annee.html'
    widget_customiz = False

    def get_context(self):
        context = super(ListeImpayesAnnee, self).get_context()
        context['liste_impayes'] = PaiementBackoffice.objects.filter(is_not_ok__exact=True)
        # context['liste_impayes'] = PaiementBackoffice.objects.filter(is_not_ok__exact=False)[0:10]
        context['years'] = AnneeUniPaiement.objects.all().order_by('-cod_anu')
        context['year'] = self.kwargs.get('year', 2014)
        return context

    @filter_hook
    def get_breadcrumb(self):
        return [{'url': self.get_admin_url('index'), 'title': 'Accueil'},
                {'url': self.get_admin_url('gestion_financiere_annee'), 'title': 'Gestion financière'},
                {'title': 'Liste des impayés'}]

    @never_cache
    def get(self, request, *args, **kwargs):
        self.widgets = self.get_widgets()
        return self.template_response(self.base_template, self.get_context())
xadmin.site.register_view(r'^liste_impayes/(?P<year>\d+)$', ListeImpayesAnnee, 'liste_impayes_annee')


class ListeInscritptionSansPaiementAnnee(views.Dashboard):
    base_template = 'duck_paiement_etudiant/liste_inscription_sans_paiement_annee.html'
    widget_customiz = False

    def get_context(self):
        context = super(ListeInscritptionSansPaiementAnnee, self).get_context()
        context['years'] = AnneeUniPaiement.objects.all().order_by('-cod_anu')
        context['inscriptions'] = InsAdmEtp.inscrits.filter(paiements__isnull=True).order_by("cod_ind__lib_nom_pat_ind")
        return context

    @filter_hook
    def get_breadcrumb(self):
        return [{'url': self.get_admin_url('index'), 'title': 'Accueil'},
                {'url': self.get_admin_url('gestion_financiere_annee'), 'title': 'Gestion financière'},
                {'title': 'Gestion financière'}]

    @never_cache
    def get(self, request, *args, **kwargs):
        self.widgets = self.get_widgets()
        context = self.get_context()
        page = request.GET.get('page')
        paginator = Paginator(context['inscriptions'], 100)
        try:
            context['inscriptions'] = paginator.page(page)

        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            context['inscriptions'] = paginator.page(1)

        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            context['inscriptions'] = paginator.page(paginator.num_pages)

        return self.template_response(self.base_template, context)

xadmin.site.register_view(r'^liste_inscription_sans_paiement/(?P<year>\d+)$',
                          ListeInscritptionSansPaiementAnnee,
                          'liste_inscription_valide_sans_paiement')


class StatistiquesBordereau(views.Dashboard):
    base_template = 'duck_paiement_etudiant/statistiques_bordereau_annee.html'
    widget_customiz = False

    def get_context(self):
        from collections import OrderedDict
        context = super(StatistiquesBordereau, self).get_context()
        year = self.kwargs.get('year', 2014)
        context['data'] = {'ordinaire': {}, 'banque': {}, 'etranger': {}, 'virement': {}}
        context['data'] = OrderedDict()
        context['data']['ordinaire'] = {'header': 'Chèque ordinaire'}
        context['data']['banque'] = {'header': 'Chèque de banque'}
        context['data']['etranger'] = {'header': 'Chèque étranger'}
        context['data']['virement'] = {'header': 'Virement'}

        context['data']['ordinaire']['nb'] = PaiementBackoffice.objects.by_year(year).filter(type='C').count()
        context['data']['etranger']['nb'] = PaiementBackoffice.objects.by_year(year).filter(type='E').count()
        context['data']['banque']['nb'] = PaiementBackoffice.objects.by_year(year).filter(type='B').count()
        context['nb_cheque_total'] = (context['data']['ordinaire']['nb'] +
                                      context['data']['etranger']['nb'] +
                                      context['data']['banque']['nb'])
        context['data']['virement']['nb'] = PaiementBackoffice.objects.by_year(year).filter(type='V').count()

        for k, v in {'ordinaire': 'C',
                     'banque': 'B',
                     'etranger': 'E',
                     'virement': 'V' }.iteritems():
            context['data'][k]['somme_totale'] = PaiementBackoffice.objects.by_year(year).filter(type=v).aggregate(Sum('somme'))['somme__sum']
            for n in range(1, 4):
                context['data'][k]['versement_{}'.format(n)] = PaiementBackoffice.objects.by_year(year).filter(type=v, num_paiement=n)\
                                                                                 .aggregate(count=Count('somme'),
                                                                                            sum=Sum('somme'))
        # Calcul combien d'étudiant à payé en 1 fois, 2 fois ou 3 fois
        result = {1:0, 2:0, 3:0}
        for x in PaiementBackoffice.objects.by_year(year).all().values('cod_ind').annotate(Count('cod_ind')):
            result[x['cod_ind__count']] += 1

        context['nb_etudiant_1_paiement'] = result[1]
        context['nb_etudiant_2_paiement'] = result[2]
        context['nb_etudiant_3_paiement'] = result[3]
        return context

    @filter_hook
    def get_breadcrumb(self):
        return [{'url': self.get_admin_url('index'), 'title': 'Accueil'},
                {'url': self.get_admin_url('gestion_financiere_annee'), 'title': 'Gestion financière'},
                {'url': self.get_admin_url('statistiques_bordereau_annee', year=self.kwargs['year']), 'title': 'Statistiques'}]

    @never_cache
    def get(self, request, *args, **kwargs):
        self.widgets = self.get_widgets()
        return self.template_response(self.base_template, self.get_context())
xadmin.site.register_view(r'^statistiques_bordereau/(?P<year>\d+)$', StatistiquesBordereau, 'statistiques_bordereau_annee')


class ImpressionBordereauAnnee(views.Dashboard):
    base_template = 'duck_paiement_etudiant/impression_bordereau_annee.html'
    widget_customiz = False

    def get_context(self):
        from collections import OrderedDict
        context = super(ImpressionBordereauAnnee, self).get_context()
        type_bordereau = self.request.GET.get('type', 'C')
        data = OrderedDict()
        data['C'] = {'title': 'Chèque ordinaire', 'is_active': False}
        data['B'] = {'title': 'Chèque de banque', 'is_active': False}
        data['E'] = {'title': 'Chèque étranger', 'is_active': False}
        data['V'] = {'title': 'Virement', 'is_active': False}
        data['A'] = {'title': 'Auditeur', 'is_active': False}
        data[type_bordereau]['is_active'] = True
        context['data'] = data
        context['year'] = self.kwargs.get('year', 2014)
        if type_bordereau != 'A':
            context['bordereaux'] = Bordereau.objects.by_year(context['year']).filter(type_paiement=type_bordereau)
        else:
            context['bordereaux'] = Bordereau.auditeur.by_year(context['year']).all()
        return context

    @filter_hook
    def get_breadcrumb(self):
        return [{'url': self.get_admin_url('index'), 'title': 'Accueil'},
                {'url': self.get_admin_url('gestion_financiere_annee'), 'title': 'Gestion financière'},
                {'title': 'Impression bordereau'}]


    @never_cache
    def get(self, request, *args, **kwargs):
        self.widgets = self.get_widgets()
        return self.template_response(self.base_template, self.get_context())
xadmin.site.register_view(r'^impression_bordereau/(?P<year>\d+)$', ImpressionBordereauAnnee, 'impression_bordereau_annee')


class ImpressionBordereau(BaseAdminView):

    def create_spreadsheet(self, bordereau):
        wb = Workbook()
        ws = wb.active

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.4
        thin_border = Border(left=Side(border_style=openpyxl.styles.borders.BORDER_THIN),
                             right=Side(border_style=openpyxl.styles.borders.BORDER_THIN),
                             top=Side(border_style=openpyxl.styles.borders.BORDER_THIN),
                             bottom=Side(border_style=openpyxl.styles.borders.BORDER_THIN))
        light_grey_fill = PatternFill(fill_type=openpyxl.styles.fills.FILL_SOLID, start_color="00FFFFFF")
        dark_grey_fill = PatternFill(fill_type=openpyxl.styles.fills.FILL_SOLID, start_color="00DDDDDD")
        arial_bold_font = Font(name="Arial", size=10, bold=True)
        arial_font = Font(name="Arial", size=10, bold=False)

        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=False)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, shrink_to_fit=False)
        right_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True, shrink_to_fit=False)

        arial_bold_style = Style(font=arial_bold_font, alignment=left_alignment)
        header_style = Style(font=arial_bold_font,
                             border=thin_border,
                             fill=dark_grey_fill,
                             alignment=center_alignment,
                             )
        arial_style = Style(font=arial_font)

        ws.merge_cells('A1:H1')
        ws['A1'].style = arial_bold_style
        ws['A1'] = u"I.E.D. Frais d'enseignement à Distance {}/{}".format(bordereau.annee.cod_anu,
                                                                          bordereau.annee.cod_anu+1)
        ws.merge_cells('A3:H3')
        ws['A3'].style = arial_bold_style
        ws['A3'] = u"MODE DE PAIEMENT: {}".format(bordereau.get_type_paiement_display())
        ws.merge_cells('A4:H4')
        ws['A4'].style = arial_style
        ws['A4'] = u"Paiement numéro {} / Bordereau numéro {}".format(bordereau.num_paiement,
                                                                      bordereau.num_bordereau)
        ## Print header
        ws['B6'].style = header_style
        ws['C6'].style = header_style
        ws['D6'].style = header_style
        ws['E6'].style = header_style
        ws['F6'].style = header_style
        ws['G6'].style = header_style
        ws['H6'].style = header_style
        ws['I6'].style = header_style
        if bordereau.type_paiement != "V":
            ws['B6'] = u"N° CHÈQUE"
            ws['C6'] = u"BANQUE"
        else: ## if Virement
            ws['B6'] = u"DATE PRÉVUE"
            ws['C6'] = u"DATE DE VIREMENT"

        ws['D6'] = u"NOM ÉTUDIANT"
        ws['E6'] = u"PRÉNOM ÉTUDIANT"
        ws['F6'] = u"CODE ÉTUDIANT"
        ws['G6'] = u"€"
        ws['H6'] = u"PAYEUR - TITULAIRE DU COMPTE"
        ws['I6'] = u"ANNULATION"

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 10
        if bordereau.type_paiement != "V":
            ws.column_dimensions['C'].width = 25
        else:
            ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 30

        row = 7
        start = 1
        if bordereau.type_bordereau == 'N':
            queryset = bordereau.paiementbackoffice_set.all().distinct().order_by('id')
        else:  # 'A'
            queryset = bordereau.paiementauditeurbackoffice_set.all().distinct().order_by('id')

        for paiement in queryset:
            cell = ws.cell(row=row, column=1)
            cell.value = start
            cell.style = Style(font=arial_bold_font,
                               border=thin_border,
                               fill=light_grey_fill if start%2 else dark_grey_fill,
                               alignment=center_alignment,)
            if bordereau.type_paiement != "V":
                ## Numero de cheque
                cell = ws.cell(row=row, column=2)
                cell.value = paiement.num_cheque
                cell.style = Style(font=arial_font,
                                   border=thin_border,
                                   fill=light_grey_fill if start%2 else dark_grey_fill,
                                   alignment=center_alignment,)
                ## Banque
                cell = ws.cell(row=row, column=3)
                cell.value = paiement.nom_banque.nom if paiement.nom_banque else "Anomalie"
                cell.style = Style(font=arial_font,
                                   border=thin_border,
                                   fill=light_grey_fill if start%2 else dark_grey_fill,
                                   alignment=left_alignment,)
            else:
                ## Date prevue
                cell = ws.cell(row=row, column=2)
                try:
                    cell.value = paiement.date.strftime('%d/%m/%Y')
                except AttributeError:
                    cell.value = ''
                cell.style = Style(font=arial_font,
                                   border=thin_border,
                                   fill=light_grey_fill if start%2 else dark_grey_fill,
                                   alignment=center_alignment,)
                ## Date Virement effectue
                cell = ws.cell(row=row, column=3)
                cell.value = (paiement.date_virement.strftime('%d/%m/%Y')
                              if paiement.date_virement else "Aucune")
                cell.style = Style(font=arial_font,
                                   border=thin_border,
                                   fill=light_grey_fill if start%2 else dark_grey_fill,
                                   alignment=center_alignment,)

            ## Nom etudiant
            cell = ws.cell(row=row, column=4)
            cell.value = paiement.etape.nom()
            cell.style = Style(font=arial_font,
                               border=thin_border,
                               fill=light_grey_fill if start%2 else dark_grey_fill,
                               alignment=left_alignment,)
            ## Prenom etudiant
            cell = ws.cell(row=row, column=5)
            cell.value = paiement.etape.prenom()
            cell.style = Style(font=arial_font,
                               border=thin_border,
                               fill=light_grey_fill if start%2 else dark_grey_fill,
                               alignment=left_alignment,)
            ## Code etudiant
            cell = ws.cell(row=row, column=6)
            cell.value = paiement.etape.cod_etu()
            cell.style = Style(font=arial_font,
                               border=thin_border,
                               fill=light_grey_fill if start%2 else dark_grey_fill,
                               alignment=left_alignment,)
            ## Somme
            cell = ws.cell(row=row, column=7)
            cell.value = paiement.somme
            cell.style = Style(font=arial_font,
                               border=thin_border,
                               fill=light_grey_fill if start%2 else dark_grey_fill,
                               alignment=right_alignment,)
            ## Payeur - Titulaire du compte
            cell = ws.cell(row=row, column=8)
            cell.value = paiement.autre_payeur
            cell.style = Style(font=arial_font,
                               border=thin_border,
                               fill=light_grey_fill if start%2 else dark_grey_fill,
                               alignment=left_alignment,)
            # Annulation
            cell = ws.cell(row=row, column=9)
            cell.value = paiement.etape.annulation()
            cell.style = Style(font=arial_font,
                               border=thin_border,
                               fill=light_grey_fill if start%2 else dark_grey_fill,
                               alignment=left_alignment,)
            row += 1
            start += 1

        ## Total
        cell = ws.cell(row=row+1, column=6)
        cell.value = "Total"
        cell.style = Style(font=arial_bold_font,
                           border=thin_border,
                           alignment=left_alignment,)
        cell = ws.cell(row=row+1, column=7)
        if bordereau.type_bordereau == 'N':
            cell.value = bordereau.all_valid().aggregate(Sum('somme'))['somme__sum']
        else:
            cell.value = bordereau.paiementauditeurbackoffice_set.aggregate(Sum('somme'))['somme__sum']

        cell.style = Style(font=arial_bold_font,
                           border=thin_border,
                           alignment=right_alignment,)
        ## Date de remise
        cell = ws.cell(row=row+3, column=6)
        cell.value = "Date de remise"
        cell.style = Style(font=arial_bold_font,
                           border=thin_border,
                           alignment=left_alignment,)
        cell = ws.cell(row=row+3, column=7)
        cell.value = bordereau.date_cloture.strftime('%d/%m/%Y') if bordereau.date_cloture else "Non clôturé"
        cell.style = Style(font=arial_bold_font,
                           border=thin_border,
                           alignment=center_alignment,)

        return wb

    def get(self, request, *args, **kwargs):
        b = Bordereau.unfiltered.get(pk=kwargs['bordereau'])
        response = HttpResponse(save_virtual_workbook(self.create_spreadsheet(b)), content_type='application/vnd.ms-excel')
        date = datetime.datetime.today().strftime('%d-%m-%Y')
        response['Content-Disposition'] = 'attachment; filename={}_{}_{}_{}_{}.xlsx'.format('bordereau', b.type_paiement,
                                                                                            b.num_paiement, b.num_bordereau,
                                                                                            date,)
        return response

xadmin.site.register_view(r'^download_bordereau_spreadsheet/(?P<bordereau>\d+)$', ImpressionBordereau, 'impression_bordereau')


class BordereauSpreadsheetView(PDFTemplateView):
    filename = "Bordereau_{}_{}_{}_{}.pdf"
    template_name = "duck_paiement_etudiant/bordereau_spreadsheet.html"
    cmd_options = {
        'orientation': 'landscape',
        'page-size': 'A4'
    }

    def get_filename(self):
        b = Bordereau.unfiltered.get(pk=self.kwargs['bordereau'])
        return self.filename.format(b.type_paiement, b.num_paiement, b.num_bordereau,
                                    datetime.datetime.today().strftime('%d-%m-%Y'))

    def get_context_data(self, **kwargs):
        b = Bordereau.unfiltered.get(pk=self.kwargs['bordereau'])
        context = super(BordereauSpreadsheetView, self).get_context_data(**kwargs)
        context['bordereau'] = b
        if b.type_bordereau == 'N':
            context['total_sum'] = b.all_valid().order_by('-id').aggregate(Sum('somme'))['somme__sum']
        else:
            context['total_sum'] = b.paiementauditeurbackoffice_set.aggregate(Sum('somme'))['somme__sum']
        return context


class PaiementInlineView(object):
    model = PaiementBackoffice
    exclude = ['cod_anu', 'cod_ind', 'cod_etp', 'cod_vrs_vet', 'num_occ_iae']
    readonly_fields = ['num_paiement', 'bordereau']
    extra = 1
    max_num = 3
    form = PaiementBackofficeForm


class PaiementAdminView(object):
    fields = [
        'get_nom', 'get_prenom',
        'get_cod_etu', 'get_adresse',
        'cod_etp', 'cod_cge',
        'get_eta_iae', 'exoneration',
        'demi_annee',
        'force_encaissement',
        'get_tarif']
    readonly_fields = [
        'get_nom', 'get_prenom',
        'get_cod_etu', 'get_adresse',
        'cod_etp', 'cod_cge',
        'get_eta_iae',
        'get_tarif']
    inlines = [PaiementInlineView]
    search_fields = ['cod_ind__cod_etu']
    hidden_menu = True

    show_bookmarks = False
    site_title = u'Dossiers financiers étudiants'
    form_layout = Layout(Container(Col('full',
                Fieldset(
                    "",
                    'get_nom', 'get_prenom',
                    'get_cod_etu', 'get_adresse',
                    'cod_etp', 'cod_cge',
                    'get_eta_iae', 'exoneration',
                    'demi_annee',
                    'force_encaissement',
                    'get_tarif'
                    , css_class="unsort no_title"), horizontal=True, span=12)
            ))
    pattern = r'^%s/%s/(?P<year>\d+)/'

    def get_kwargs_url(self, instance=None):
        return {'year': 2014}

    def queryset(self):
        return InsAdmEtpPaiement.objects.filter(cod_anu=self.kwargs['year'], cod_cge='IED', tem_iae_prm='O')

    # def get_tarif(self, obj):
    #     print obj.get_tarif()
    #     return obj.get_tarif()

    @filter_hook
    def get_breadcrumb(self):
        breadcrumb = super(PaiementAdminView, self).get_breadcrumb()
        breadcrumb = [breadcrumb[0]] + [{'url': self.get_admin_url('gestion_financiere_annee'), 'title': 'Gestion financière'}] + breadcrumb[1:]
        return breadcrumb

    def get_media(self, *args, **kwargs):
        media = super(PaiementAdminView, self).get_media(*args, **kwargs)
        m = Media()
        m.add_js(['paiement_etudiant/js/paiement_etudiant.js'])
        return media+m

    def get_nom(self, obj):
        return obj.cod_ind.lib_nom_pat_ind
    get_nom.short_description = 'Nom'
    get_nom.allow_tags = True

    def get_prenom(self, obj):
        return '{}'.format(obj.cod_ind.lib_pr1_ind)
    get_prenom.short_description = 'Prenom'
    get_prenom.allow_tags = True

    def get_cod_etu(self, obj):
        return '{}'.format(obj.cod_ind.cod_etu)
    get_cod_etu.short_description = 'Code étudiant'
    get_cod_etu.allow_tags = True

    def get_cod_opi(self, obj):
        return '{}'.format(obj.cod_ind.cod_ind_opi)
    get_cod_opi.short_description = 'Code opi'
    get_cod_opi.allow_tags = True

    def get_adresse(self, obj):
        return '{}'.format(obj.cod_ind.get_full_adresse(obj.cod_anu.cod_anu))
    get_adresse.short_description = 'Adresse'
    get_adresse.allow_tags = True

    def get_eta_iae(self, obj):
        return '{}'.format(obj.annulation())
    get_eta_iae.short_description = 'Etat de l\'inscription administrative'
    get_eta_iae.allow_tags = True



class BanqueAdmin(object):
    search_fields = ['nom']


class MyFilter(NumberFieldListFilter):
    def do_filte(self, queryset):
        return super(MyFilter, self).do_filte(queryset)


class BordereauAdmin(object):
    fields = ('num_bordereau', 'num_paiement', 'cloture',
                       'get_annee', 'get_type_paiement',
                       'get_type_bordereau', 'date_cloture',
                       'envoi_mail',
                       'get_total_sum', 'get_nb_cheque_total', 'comment')
    readonly_fields = ('num_bordereau', 'num_paiement',
                       'get_annee', 'get_type_paiement',
                       'get_type_bordereau', 'date_cloture',
                       'envoi_mail',
                       'get_total_sum', 'get_nb_cheque_total')
    hidden_menu = True
    list_display = ('__str__', 'num_paiement', 'type_paiement', 'date_cloture', 'cloture',  'comment')
    list_filter = ('type_paiement', ('num_paiement', MyFilter), 'type_bordereau')
    form_layout = Layout(Container(Col('full',
                Fieldset(
                    '',
                    "id",
                    'num_bordereau', 'num_paiement',
                    'get_annee', 'get_type_paiement',
                    'envoi_mail',
                    'get_type_bordereau', 'date_cloture',
                    'get_total_sum',
                    'get_nb_cheque_total',
                    'cloture', 'comment',
                    css_class="unsort no_title"), horizontal=True, span=12)
            ))
    pattern = r'^%s/%s/(?P<year>\d+)/'

    def get_kwargs_url(self, instance=None):
        return {'year': 2014}

    @filter_hook
    def get_breadcrumb(self):
        return [{'url': self.get_admin_url('index'), 'title': 'Accueil'},
                {'url': self.get_admin_url('gestion_financiere_annee'), 'title': 'Gestion financière'},
                {'title': 'Bordereaux'}]

    def queryset(self):
        queryset = super(BordereauAdmin, self).queryset()
        year = self.kwargs.get('year', 2014)
        return queryset.filter(annee__cod_anu=year)

    def get_annee(self, obj):
        return '{}'.format(obj.annee)
    get_annee.short_description = "Année"

    def get_type_paiement(self, obj):
        return obj.get_type_paiement_display()
    get_type_paiement.short_description = "Type de paiement"

    def get_type_bordereau(self, obj):
        return obj.get_type_bordereau_display()
    get_type_bordereau.short_description = "Type de bordereau"

    def get_total_sum(self, obj):
        return '{}'.format(obj.total_sum())
    get_total_sum.short_description = 'Somme totale'
    get_total_sum.allow_tags = True

    def get_nb_cheque_total(self, obj):
        return '{}'.format(obj.nb_cheque_total())
    get_nb_cheque_total.short_description = 'Nombre de chèque(s)'
    get_nb_cheque_total.allow_tags = True


class PaiementAuditeurBackofficeInlineAdmin(object):
    model = PaiementAuditeurBackoffice
    exclude = ['cod_anu', 'cod_etp']
    readonly_fields = ['num_paiement', 'bordereau']
    extra = 1
    max_num = 3
    form = PaiementBackofficeForm


class AuditeurLibreApogeeAdmin(object):
    inlines = [PaiementAuditeurBackofficeInlineAdmin]

    fields = [
        'get_nom', 'get_prenom',
        'get_cod_ied', 'get_adresse',
        'get_tarif']

    readonly_fields = [
        'get_nom', 'get_prenom',
        'get_cod_ied', 'get_adresse',
        'get_tarif']

    search_fields = ['etape__code_ied', 'etape__last_name', 'etape__first_name']
    hidden_menu = True

    show_bookmarks = False
    site_title = u'Dossiers financiers auditeur libre'
    form_layout = Layout(Container(Col('full',
                Fieldset(
                    "",
                    'get_nom', 'get_prenom',
                    'get_cod_ied', 'get_adresse',
                    'cod_etp',
                    'get_tarif'
                    , css_class="unsort no_title"), horizontal=True, span=12)
            ))
    pattern = r'^%s/%s/(?P<year>\d+)/'

    def get_kwargs_url(self, instance=None):
        return {'year': 2014}

    def queryset(self):
        return AuditeurLibreApogee.objects.filter(annee__cod_anu=self.kwargs['year'])

    def get_tarif(self, obj):
        return ''

    @filter_hook
    def get_breadcrumb(self):
        breadcrumb = super(AuditeurLibreApogeeAdmin, self).get_breadcrumb()
        breadcrumb = [breadcrumb[0]] + [{'url': self.get_admin_url('gestion_financiere_annee'), 'title': 'Gestion financière'}] + breadcrumb[1:]
        return breadcrumb

    def get_media(self, *args, **kwargs):
        media = super(AuditeurLibreApogeeAdmin, self).get_media(*args, **kwargs)
        m = Media()
        m.add_js(['paiement_etudiant/js/paiement_etudiant.js'])
        return media+m

    def get_nom(self, obj):
        return obj.last_name
    get_nom.short_description = 'Nom'
    get_nom.allow_tags = True

    def get_prenom(self, obj):
        return '{}'.format(obj.first_name)
    get_prenom.short_description = 'Prenom'
    get_prenom.allow_tags = True

    def get_cod_ied(self, obj):
        return '{}'.format(obj.code_ied)
    get_cod_ied.short_description = 'Code ied'
    get_cod_ied.allow_tags = True

    def get_adresse(self, obj):
        return '{}'.format(obj.address)
    get_adresse.short_description = 'Adresse'
    get_adresse.allow_tags = True


xadmin.site.register(InsAdmEtpPaiement, PaiementAdminView)
xadmin.site.register(Banque, BanqueAdmin)
xadmin.site.register(Bordereau, BordereauAdmin)
xadmin.site.register(SettingEtapePaiement)
xadmin.site.register(AuditeurLibreApogee, AuditeurLibreApogeeAdmin)